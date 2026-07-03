#!/usr/bin/env python3
"""
ACVSTMC Dashboard Data Generator
Reads roster Excel + attendance log → data.json for the dashboard website.
"""
import json, re, os, sys
from datetime import datetime, timezone, timedelta
from collections import defaultdict
import openpyxl

# === CONFIG ===
HERE = os.path.dirname(os.path.abspath(__file__))
ROSTER_PATH = '/home/kenny/ACVSTMC_Roster_Complete_July2026.xlsx'
LOG_PATH = '/home/kenny/.hermes/profiles/housing-admin/whatsapp/attendance-log.json'
LID_DIR = '/home/kenny/.hermes/profiles/housing-admin/whatsapp/session'
CONTACT_PATH = '/home/kenny/ACVSTMC_所有邨_更表總覽_202607.xlsx'
OUTPUT_PATH = os.path.join(HERE, 'data.json')

HK_TZ = timedelta(hours=8)

# === STEP 0: Phone mapping from contact list ===
def load_phone_map():
    """Load phone → (staff_no, chinese_name, english_name, estate, shift_pattern)"""
    wb = openpyxl.load_workbook(CONTACT_PATH, data_only=True)
    phone_map = {}
    estate_worker_map = defaultdict(list)  # estate → [(chinese_name, staff_no, shift_pattern)]
    
    for sname in wb.sheetnames:
        ws = wb[sname]
        for row in ws.iter_rows(min_row=2, values_only=True):
            sno = str(row[0]).strip() if row[0] else ''
            if not sno or not re.match(r'W\d+', sno):
                continue
            ename = str(row[1] or '').strip()
            cname = str(row[2] or '').strip()
            phone_raw = str(row[4] or '').strip() if len(row) > 4 else ''
            phone = re.sub(r'[\s\-\(\)]', '', phone_raw)  # strip spaces/hyphens
            shift = str(row[5] or '').strip() if len(row) > 5 else ''
            remarks = str(row[6] or '').strip() if len(row) > 6 else ''
            
            if phone:
                phone_map[phone] = (sno, cname, ename, sname, shift)
            
            # Also map by estate
            estate_worker_map[sname].append({
                'staff_no': sno,
                'chinese_name': cname,
                'english_name': ename,
                'shift_pattern': shift,
                'remarks': remarks,
            })
    
    return phone_map, estate_worker_map, wb

# === STEP 1: Load lid → phone mapping ===
def load_lid_map():
    """Load lid → phone from Baileys session files"""
    lid_map = {}
    if not os.path.isdir(LID_DIR):
        return lid_map
    for fname in os.listdir(LID_DIR):
        if fname.startswith('lid-mapping-') and '_reverse' in fname:
            lid = fname.replace('lid-mapping-', '').replace('_reverse.json', '')
            path = os.path.join(LID_DIR, fname)
            try:
                with open(path) as f:
                    phone = f.read().strip().strip('"')
                lid_map[lid] = phone
            except:
                pass
    return lid_map

# === STEP 2: Load attendance log ===
def load_attendance():
    """Load attendance log, return list of records"""
    try:
        with open(LOG_PATH) as f:
            data = json.load(f)
        return data
    except:
        return []

# === STEP 3: Load roster Excel ===
def load_roster():
    """Load daily roster from Roster_Complete Excel"""
    wb = openpyxl.load_workbook(ROSTER_PATH, data_only=True)
    roster = {}  # estate_name → {worker_name → {day: shift_code}}
    estate_info = {}  # estate_name → shift_definitions
    
    for sname in wb.sheetnames:
        if sname == '目錄' or sname == '司機':
            continue
        ws = wb[sname]
        
        # Row 1: title
        # Row 2: shift definitions
        shift_defs = str(ws.cell(2, 1).value or '')
        # Row 5: header with day numbers 1-31
        # Row 6: day of week
        # Row 7+: worker data
        
        workers = {}
        for r in range(7, ws.max_row + 1):
            name = str(ws.cell(r, 1).value or '').strip()
            if not name or name.isdigit() or len(name) < 2:
                continue
            if name in ('圖例', '備註', 'ACVSTMC', '') or any(kw in name for kw in ('Hsin', 'Generated')):
                continue
            
            shifts = {}
            for c in range(2, min(33, ws.max_column + 1)):  # cols B=2 to AF=32 = days 1-31
                val = str(ws.cell(r, c).value or '').strip()
                day_num = c - 1
                if val:
                    shifts[day_num] = val
            if shifts:
                workers[name] = shifts
        
        if workers:
            roster[sname] = workers
            estate_info[sname] = shift_defs
    
    return roster, estate_info, wb

# === STEP 4: Estate name mapping ===
ROSTER_TO_ESTATE = {
    '房署總部1,2座': '房委會(1,2座)',
    '房署總部3座': '房委會(3座)',
    '房委會(4座)': '房委會(4座)',
    '華富1': '華富(一)',
    '迎東': '迎東街市',
    '海達': '海達街市',
    '水泉澳廣場': '水泉澳',
    '石門': '石門',
    '美田商場': '美田邨',
    '皇后山': '皇后山',
    '麗晶廣場(駿洋商場)': '駿洋邨',
    '長青商場(晴朗商場)': '晴朗商場',
    '海麗商場': '海麗商場',
    '石硤尾邨': '石硤尾',
    '滿東市場': '滿東街市',
}

# Known estate names in attendance log
ESTATE_NAMES_IN_LOG = list(set(ROSTER_TO_ESTATE.values()))

# === STEP 5: Build the daily data ===
def get_today_date():
    """Get today's HK date"""
    utc_now = datetime.now(timezone.utc)
    hk_now = utc_now + HK_TZ
    return hk_now.strftime('%Y-%m-%d'), hk_now.day

def build_dashboard_data():
    phone_map, estate_worker_map, contact_wb = load_phone_map()
    lid_map = load_lid_map()
    records = load_attendance()
    roster, estate_info, roster_wb = load_roster()
    
    today_str, today_day = get_today_date()
    
    # Build worker mapping from lid → worker info
    # First invert phone_map: phone → (staff_no, chinese_name, estate)
    phone_to_worker = {}
    for phone, (sno, cname, ename, estate, shift) in phone_map.items():
        phone_to_worker[phone] = {'staff_no': sno, 'chinese_name': cname, 'english_name': ename, 'estate': estate, 'shift_pattern': shift}
    
    # Also build name → estate mapping from contacts
    name_to_estate = {}
    for phone, (sno, cname, ename, estate, shift) in phone_map.items():
        if cname:
            name_to_estate[cname] = estate
        if ename:
            name_to_estate[ename.lower()] = estate
    
    # Group attendance records by estate for today
    today_records = [r for r in records if r.get('date') == today_str and r.get('hasPhoto')]
    
    # Group photos by sender lid for today
    sender_photos = defaultdict(list)
    for r in today_records:
        sender_photos[r['sender']].append(r)
    
    # Map sender lids to worker info
    lid_worker = {}  # lid → worker info
    for lid in sender_photos:
        phone = lid_map.get(lid, '')
        if not phone:
            continue
        phone_clean = phone.replace('852', '', 1) if phone.startswith('852') else phone
        worker = phone_to_worker.get(phone_clean) or phone_to_worker.get(phone)
        if worker:
            lid_worker[lid] = worker
    
    # Build estate data for today
    estates_data = []
    
    # Process each estate from roster
    for roster_name, workers in roster.items():
        estate_name = ROSTER_TO_ESTATE.get(roster_name, roster_name)
        
        worker_list = []
        for wname, shifts in workers.items():
            today_shift = shifts.get(today_day, '')
            
            # Find worker in contact list by name
            worker_info = None
            # Try to match by Chinese name
            for phone, info in phone_to_worker.items():
                if info['chinese_name'] == wname or info['english_name'].lower() == wname.lower():
                    worker_info = info
                    break
            
            cname = (worker_info or {}).get('chinese_name', wname)
            
            # Find attendance for this worker
            clock_in = ''
            clock_out = ''
            for lid, winfo in lid_worker.items():
                if winfo.get('staff_no') == (worker_info or {}).get('staff_no'):
                    photos = sender_photos[lid]
                    times = sorted([p['time'] for p in photos], key=lambda t: t.replace(':', ''))
                    if times:
                        # Check time to determine in/out
                        # Morning: <= 13:00 → clock-in; Afternoon > 13:00 → clock-out
                        in_times = [t for t in times if t <= '13:00']
                        out_times = [t for t in times if t > '13:00']
                        if in_times:
                            clock_in = in_times[0]
                        if out_times:
                            clock_out = out_times[-1]
                    break
            
            # Get monthly shift data for calendar view
            monthly_shifts = []
            for d in range(1, 32):
                monthly_shifts.append(shifts.get(d, ''))
            
            sno = (worker_info or {}).get('staff_no', '')
            
            worker_list.append({
                'name': wname,
                'chineseName': cname,
                'staffNo': sno,
                'shift': today_shift if today_shift else '—',
                'clockIn': clock_in,
                'clockOut': clock_out,
                'monthlyShifts': monthly_shifts,
            })
        
        estates_data.append({
            'name': estate_name,
            'rosterName': roster_name,
            'shiftDefs': estate_info.get(roster_name, ''),
            'workers': worker_list,
        })
    
    # Also include estates that have attendance but no roster (like 房委會(4座))
    attended_estates = set(r['estate'] for r in today_records)
    roster_estates = set(ROSTER_TO_ESTATE.values())
    
    for estate in attended_estates:
        if estate not in roster_estates:
            # Find workers from this estate in contact list
            workers_in_estate = estate_worker_map.get(estate, [])
            worker_list = []
            for w in workers_in_estate:
                # Check if they have attendance today
                clock_in = ''
                clock_out = ''
                for lid, winfo in lid_worker.items():
                    if winfo.get('staff_no') == w['staff_no']:
                        photos = sender_photos[lid]
                        times = sorted([p['time'] for p in photos], key=lambda t: t.replace(':', ''))
                        if times:
                            in_times = [t for t in times if t <= '13:00']
                            out_times = [t for t in times if t > '13:00']
                            if in_times:
                                clock_in = in_times[0]
                            if out_times:
                                clock_out = out_times[-1]
                        break
                
                worker_list.append({
                    'name': w['english_name'],
                    'chineseName': w['chinese_name'],
                    'staffNo': w['staff_no'],
                    'shift': w['shift_pattern'][:5] if w['shift_pattern'] else '?',
                    'clockIn': clock_in,
                    'clockOut': clock_out,
                    'monthlyShifts': [],
                })
            
            estates_data.append({
                'name': estate,
                'rosterName': '',
                'shiftDefs': '',
                'workers': worker_list,
            })
    
    # Sort estates by name
    estates_data.sort(key=lambda e: e['name'])
    
    # Count stats
    total_workers = 0
    clocked_in = 0  # has clock-in
    clocked_out = 0  # has clock-out
    no_photo = 0  # no photos at all
    for estate in estates_data:
        for w in estate['workers']:
            if w['shift'] in ('H', 'P', 'S', '—'):
                continue  # rest day or unknown
            total_workers += 1
            if w['clockIn'] and w['clockOut']:
                clocked_in += 1
                clocked_out += 1
            elif w['clockIn']:
                clocked_in += 1
            elif w['clockOut']:
                clocked_out += 1
            else:
                no_photo += 1
    
    present_any = clocked_in + clocked_out - (clocked_in if clocked_in < clocked_out else clocked_out)
    # Actually simpler: has_any_photo = total_workers - no_photo
    has_any_photo = total_workers - no_photo
    
    output = {
        'lastUpdated': datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC'),
        'date': today_str,
        'totalWorkers': total_workers,
        'clockedIn': clocked_in,
        'clockedOut': clocked_out,
        'noPhoto': no_photo,
        'hasPhoto': has_any_photo,
        'attendanceRate': round(has_any_photo / total_workers * 100, 1) if total_workers > 0 else 0,
        'estates': [{
            'name': e['name'],
            'shiftDefs': e['shiftDefs'],
            'workers': e['workers'],
        } for e in estates_data],
        'days': {
            today_str: {
                'date': today_str,
                'estates': [{
                    'name': e['name'],
                    'shiftDefs': e['shiftDefs'],
                    'workers': e['workers'],
                } for e in estates_data],
            }
        },
        'estateSummary': sorted([{
            'name': e['name'],
            'total': len([w for w in e['workers'] if w['shift'] not in ('H', 'P', 'S', '—')]),
            'clockedIn': len([w for w in e['workers'] if w['clockIn'] and w['shift'] not in ('H', 'P', 'S', '—')]),
            'clockedOut': len([w for w in e['workers'] if w['clockOut'] and w['shift'] not in ('H', 'P', 'S', '—')]),
            'noPhoto': len([w for w in e['workers'] if not w['clockIn'] and not w['clockOut'] and w['shift'] not in ('H', 'P', 'S', '—')]),
            'rest': len([w for w in e['workers'] if w['shift'] in ('H', 'P', 'S', '—')]),
        } for e in estates_data], key=lambda e: e['name']),
    }
    
    with open(OUTPUT_PATH, 'w') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    
    print(f"✅ Dashboard data generated: {OUTPUT_PATH}")
    print(f"   Date: {today_str}")
    print(f"   Total workers: {total_workers}")
    print(f"   Clocked-in: {clocked_in} | Clocked-out: {clocked_out} | No photo: {no_photo}")
    print(f"   Has any photo: {has_any_photo} / {total_workers}")
    print(f"   Attendance rate: {output['attendanceRate']}%")
    print(f"   Estates in data: {len(estates_data)}")
    return output

if __name__ == '__main__':
    build_dashboard_data()
